import hashlib
import os
import base64
import getpass
import json
from pathlib import Path
import argparse

# Use cryptography library for Ed25519 operations (more maintained than ed25519)
from cryptography.hazmat.primitives.asymmetric import ed25519
from cryptography.hazmat.primitives import serialization

# For BIP-39 mnemonic support
from mnemonic import Mnemonic

def generate_secure_seed():
    """Generate a cryptographically secure random seed"""
    return os.urandom(32)  # Use OS-provided cryptographically secure random data

def generate_seed_from_passphrase(passphrase):
    """Generate a seed from a user-provided passphrase"""
    if not passphrase:
        raise ValueError("Passphrase cannot be empty")
    # Use a strong KDF with salt for passphrase-based seed generation
    salt = b"sui-wallet-generator"  # Fixed salt, could be made configurable
    return hashlib.pbkdf2_hmac('sha256', passphrase.encode('utf-8'), salt, 100000)

def generate_mnemonic_from_seed(seed):
    """Generate a BIP-39 mnemonic from a seed"""
    # Ensure the seed is of correct entropy size (128-256 bits)
    # BIP-39 needs specific entropy sizes, so we'll use 256 bits (32 bytes)
    mnemo = Mnemonic("english")
    # Convert seed to binary entropy 
    entropy = hashlib.sha256(seed).digest()
    # Generate mnemonic from entropy
    return mnemo.to_mnemonic(entropy)

def generate_address(seed):
    """Generate a SUI address from a 32-byte seed"""
    if not isinstance(seed, bytes) or len(seed) != 32:
        raise ValueError("Seed must be a 32-byte value")

    # Generate mnemonic
    mnemonic = generate_mnemonic_from_seed(seed)
    
    # Generate the private key from seed
    private_key = ed25519.Ed25519PrivateKey.from_private_bytes(seed)
    
    # Get the public key
    public_key = private_key.public_key()
    
    # Get the raw bytes
    private_key_bytes = private_key.private_bytes(
        encoding=serialization.Encoding.Raw,
        format=serialization.PrivateFormat.Raw,
        encryption_algorithm=serialization.NoEncryption()
    )
    
    public_key_bytes = public_key.public_bytes(
        encoding=serialization.Encoding.Raw,
        format=serialization.PublicFormat.Raw
    )
    
    # Convert to hex
    skey_hex = private_key_bytes.hex()
    vkey_hex = public_key_bytes.hex()
    
    # Scheme flag byte (0x00 for Ed25519)
    scheme_flag_byte = b'\x00'
    
    # Concatenate the scheme flag byte and the public key bytes
    data_to_hash = scheme_flag_byte + public_key_bytes
    
    # Hash the concatenated data using BLAKE2b
    address_hash = hashlib.blake2b(data_to_hash, digest_size=32).digest()
    
    # Format the address
    address = "0x" + address_hash.hex()
    
    return (skey_hex, vkey_hex, address, mnemonic)

def save_wallet_info(wallet_info, batch_mode=False, excel_export=False):
    """Save wallet info to the specified files"""
    # Original saving code remains the same
    if batch_mode:
        # Create files or clear them if they already exist
        with open("wallet.txt", "w") as f:
            pass
        with open("privatekeys.txt", "w") as f:
            pass
        with open("mnemonic.txt", "w") as f:
            pass
        with open("SuiWallet.txt", "w") as f:
            f.write("address,privatekey,mnemonic\n")  # Header
        
        # Write all wallets to each file
        for wallet in wallet_info:
            # Address to wallet.txt
            with open("wallet.txt", "a") as f:
                f.write(f"{wallet['address']}\n")
            
            # Private key to privatekeys.txt
            with open("privatekeys.txt", "a") as f:
                f.write(f"{wallet['private_key']}\n")
            
            # Mnemonic to mnemonic.txt
            with open("mnemonic.txt", "a") as f:
                f.write(f"{wallet['mnemonic']}\n")
            
            # All info to SuiWallet.txt
            with open("SuiWallet.txt", "a") as f:
                f.write(f"{wallet['address']},{wallet['private_key']},{wallet['mnemonic']}\n")
        
        print(f"Batch of {len(wallet_info)} wallets created!")
        print(f"- Addresses saved to wallet.txt")
        print(f"- Private keys saved to privatekeys.txt")
        print(f"- Mnemonics saved to mnemonic.txt")
        print(f"- Complete wallet info saved to SuiWallet.txt")
        
        # Export to Excel if requested
        if excel_export:
            export_to_excel(wallet_info, batch_mode=True)
        
        return
    else:
        # Single wallet - append to existing files
        wallet = wallet_info
        
        # Address to wallet.txt
        with open("wallet.txt", "a") as f:
            f.write(f"{wallet['address']}\n")
        
        # Private key to privatekeys.txt
        with open("privatekeys.txt", "a") as f:
            f.write(f"{wallet['private_key']}\n")
        
        # Mnemonic to mnemonic.txt
        with open("mnemonic.txt", "a") as f:
            f.write(f"{wallet['mnemonic']}\n")
        
        # All info to SuiWallet.txt
        # Check if file exists and add header if needed
        if not os.path.exists("SuiWallet.txt") or os.path.getsize("SuiWallet.txt") == 0:
            with open("SuiWallet.txt", "w") as f:
                f.write("address,privatekey,mnemonic\n")  # Header
        
        with open("SuiWallet.txt", "a") as f:
            f.write(f"{wallet['address']},{wallet['private_key']},{wallet['mnemonic']}\n")
        
        print("Wallet information saved:")
        print("- Address added to wallet.txt")
        print("- Private key added to privatekeys.txt")
        print("- Mnemonic added to mnemonic.txt")
        print("- Complete wallet info added to SuiWallet.txt")
        
        # Export to Excel if requested
        if excel_export:
            export_to_excel(wallet, batch_mode=False)

def generate_multiple_wallets(count, method="random", passphrase=None):
    """Generate multiple wallets at once"""
    wallets = []
    
    for i in range(count):
        try:
            if method == "random":
                seed = generate_secure_seed()
            elif method == "passphrase" and passphrase:
                # For multiple wallets with passphrase, add counter to make them unique
                modified_passphrase = f"{passphrase}-{i+1}"
                seed = generate_seed_from_passphrase(modified_passphrase)
            else:
                raise ValueError("Invalid method or missing passphrase")
                
            skey_hex, vkey_hex, address, mnemonic = generate_address(seed)
            
            wallet_info = {
                "address": address,
                "public_key": vkey_hex,
                "private_key": skey_hex,
                "mnemonic": mnemonic
            }
            
            wallets.append(wallet_info)
            print(f"Generated wallet {i+1}/{count}: {address}")
            
        except Exception as e:
            print(f"Error generating wallet {i+1}: {e}")
    
    return wallets

def export_to_excel(wallet_info, batch_mode=False):
    """Export wallet information to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        print("Exporting wallets to Excel...")
        
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "SUI Wallets"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2B5DCD", end_color="2B5DCD", fill_type="solid")  # SUI blue color
        
        # Add header
        headers = ["Address", "Private Key", "Mnemonic Phrase"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add wallet data
        wallets_to_add = wallet_info if batch_mode else [wallet_info]
        
        for row_num, wallet in enumerate(wallets_to_add, 2):  # Start from row 2 (after header)
            ws.cell(row=row_num, column=1).value = wallet['address']
            ws.cell(row=row_num, column=2).value = wallet['private_key']
            ws.cell(row=row_num, column=3).value = wallet['mnemonic']
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 45  # Address
        ws.column_dimensions['B'].width = 70  # Private Key
        ws.column_dimensions['C'].width = 90  # Mnemonic
        
        # Add security warning
        warning_row = len(wallets_to_add) + 3
        warning_cell = ws.cell(row=warning_row, column=1)
        warning_cell.value = "⚠️ SECURITY WARNING: Keep your private keys and mnemonics secure! Never share them with anyone!"
        warning_cell.font = Font(name='Arial', size=12, bold=True, color="FF0000")
        ws.merge_cells(start_row=warning_row, start_column=1, end_row=warning_row, end_column=3)
        
        # Save the workbook
        excel_filename = "SuiWallets.xlsx"
        wb.save(excel_filename)
        print(f"Wallet information exported to {excel_filename}")
        return True
    
    except ImportError:
        print("Excel export requires openpyxl package. Install with: pip install openpyxl")
        return False
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        return False

def main():
    # Check if script is run with command line arguments
    parser = argparse.ArgumentParser(description="SUI Wallet Generator")
    parser.add_argument('--count', type=int, help='Number of wallets to generate (batch mode)')
    parser.add_argument('--random', action='store_true', help='Use random seed generation (for batch mode)')
    parser.add_argument('--passphrase', type=str, help='Use passphrase seed generation (for batch mode)')
    parser.add_argument('--excel', action='store_true', help='Export wallet information to Excel file')
    
    args = parser.parse_args()
    
    # If arguments are provided, run in batch mode
    if args.count and args.count > 0 and (args.random or args.passphrase):
        print(f"\n=== SUI Wallet Generator - Batch Mode ===")
        print(f"Generating {args.count} wallets...")
        
        method = "random" if args.random else "passphrase"
        passphrase = args.passphrase if method == "passphrase" else None
        
        if method == "passphrase" and not passphrase:
            passphrase = getpass.getpass("Enter a base passphrase (will not be displayed): ")
        
        wallets = generate_multiple_wallets(args.count, method, passphrase)
        excel_export = args.excel
        if excel_export:
            print("Excel export enabled via command line")
        save_wallet_info(wallets, batch_mode=True, excel_export=excel_export)
        return

    # Interactive mode
    print("\n=== SUI Wallet Generator ===")
    print("SECURITY WARNING: Keep your private key and seed phrase secure!")
    print("Never share them with anyone or store them in plain text.\n")
    
    # Ask if user wants to generate multiple wallets
    multi_choice = input("Do you want to generate multiple wallets? (yes/no): ").lower()
    
    if multi_choice == "yes":
        try:
            count = int(input("How many wallets do you want to generate? "))
            if count <= 0:
                print("Number must be positive.")
                return
                
            method_choice = input("Select seed generation method:\n1. Random (most secure)\n2. Passphrase-based\nChoice (1/2): ")
            
            method = ""
            passphrase = None
            
            if method_choice == "1":
                method = "random"
            elif method_choice == "2":
                method = "passphrase"
                passphrase = getpass.getpass("Enter a base passphrase (will not be displayed): ")
                confirmation = getpass.getpass("Confirm passphrase: ")
                
                if passphrase != confirmation:
                    print("Passphrases don't match!")
                    return
            else:
                print("Invalid choice.")
                return
                
            wallets = generate_multiple_wallets(count, method, passphrase)
            
            # Ask about Excel export in interactive mode
            excel_choice = input("Do you want to export wallet data to Excel? (yes/no): ").lower()
            excel_export = excel_choice == "yes"
            
            save_wallet_info(wallets, batch_mode=True, excel_export=excel_export)
            return
            
        except ValueError:
            print("Please enter a valid number.")
            return
    
    # Single wallet generation
    method = input("Select seed generation method:\n1. Random (most secure)\n2. Passphrase\nChoice (1/2): ")
    
    seed = None
    if method == "1":
        seed = generate_secure_seed()
        seed_backup = base64.b64encode(seed).decode('utf-8')
        print(f"\nYour seed (base64 encoded for backup):\n{seed_backup}")
        print("\nWARNING: Store this seed securely. It gives access to your wallet!")
    elif method == "2":
        passphrase = getpass.getpass("Enter a strong passphrase (will not be displayed): ")
        confirmation = getpass.getpass("Confirm passphrase: ")
        
        if passphrase != confirmation:
            print("Passphrases don't match!")
            return
        
        seed = generate_seed_from_passphrase(passphrase)
    else:
        print("Invalid choice.")
        return
    
    try:
        skey_hex, vkey_hex, address, mnemonic = generate_address(seed)
        
        print("\n=== Wallet Information ===")
        print(f"SUI Address: {address}")
        print(f"Public Key: {vkey_hex}")
        
        show_sensitive = input("\nDo you want to display sensitive information (mnemonic/private key)? (yes/no): ").lower()
        if show_sensitive == "yes":
            print(f"\nMnemonic Phrase: {mnemonic}")
            print(f"Private Key: {skey_hex}")
        
        wallet_info = {
            "address": address,
            "public_key": vkey_hex,
            "private_key": skey_hex,
            "mnemonic": mnemonic
        }
        
        save_option = input("\nDo you want to save the wallet information? (yes/no): ").lower()
        if save_option == "yes":
            excel_option = input("Export to Excel file? (yes/no): ").lower()
            excel_export = excel_option == "yes"
            save_wallet_info(wallet_info, excel_export=excel_export)
        
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()